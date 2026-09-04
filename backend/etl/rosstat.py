"""Ряды перевозок грузов автомобильным транспортом.

Росстат публикует показатели раздела «Транспорт» книгами Excel: строка —
субъект федерации, колонка — год, лист — круг перевозчиков. Устойчивого адреса
машиночитаемой выгрузки у публикации нет, поэтому сами книги входят в поставку
проекта, а загрузчик читает их напрямую. Числа в базе получены из
первоисточника и сверяются с ним построчно, а не переписаны от руки
в промежуточный формат.

Из книг берутся четыре ряда: перевозка грузов и грузооборот, каждый —
по всем перевозчикам и отдельно по перевозкам на коммерческой основе.
Ряды сводятся в одну запись на год и территорию: перевозка и грузооборот
описывают одну и ту же работу транспорта с разных сторон, и вместе они дают
среднее расстояние перевозки — величину, которой ни один из рядов
по отдельности не содержит.

Территории берутся четыре: город Москва как предмет системы, Московская
область как смежная территория складского пояса, Центральный федеральный
округ и Россия — как основание для сравнения. Все ряды помечаются измеренными:
это опубликованные наблюдения, а не расчёт системы.
"""

from __future__ import annotations

import logging
import re
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from core.choices import (
    DataOrigin,
    FlowDirection,
    FlowScope,
    PeriodType,
    SourceType,
    UpdateFrequency,
)
from core.models import DataSource, FreightFlowStat

from .pipeline import Candidate, Context, Extract, ModelPipeline, RunReport
from .quality import Check, not_negative, required, within
from .reference import ReferenceError, read, reference_dir

logger = logging.getLogger("freightflow.etl")

#: Код источника в справочнике системы.
SOURCE_CODE = "rosstat"

#: Описание рядов и их происхождения.
MANIFEST_FILE = "freight_statistics.json"

#: Подкаталог с книгами публикации.
BOOKS_DIR = "rosstat"

#: Первый год, за который публикуются ряды. Значение служит границей
#: правдоподобия: год вне этих пределов означает, что разобрана не та строка.
FIRST_YEAR = 1990

#: Сноска в подписи строки или колонки: «Российская Федерация1)», «20253)».
#: Номер сноски — одна цифра: в публикации их не больше трёх, а допущение
#: нескольких цифр съело бы у года последние разряды.
_FOOTNOTE_RE = re.compile(r"\s*\d\)\s*$")

#: Год в заголовке колонки.
_YEAR_RE = re.compile(r"^(\d{4})")


class RosstatError(ReferenceError):
    """Книга публикации отсутствует или её разметка изменилась."""


@dataclass(frozen=True)
class SeriesSpec:
    """Один ряд публикации: книга, лист, показатель и единица измерения."""

    file: str
    sheet: str
    measure: str
    scope: str
    unit: str
    factor: int
    title: str


@dataclass
class Observation:
    """Наблюдение: территория, год и значения показателей."""

    territory_code: str
    territory: str
    year: int
    values: dict[str, float]
    sources: dict[str, str]

    @property
    def key(self) -> str:
        return f"{self.territory_code}:{self.year}"


def strip_footnote(text: Any) -> str:
    """Убрать номер сноски из подписи строки или колонки."""
    return _FOOTNOTE_RE.sub("", str(text or "")).strip()


def books_dir() -> Path:
    return reference_dir() / BOOKS_DIR


def ensure_source() -> DataSource:
    """Справочная запись об источнике статистических рядов."""
    manifest = read(MANIFEST_FILE)
    source, _ = DataSource.objects.update_or_create(
        code=SOURCE_CODE,
        defaults={
            "name": "Росстат — перевозки грузов автомобильным транспортом",
            "source_type": SourceType.OPEN_DATA,
            "url": manifest.source.get("url", ""),
            "update_frequency": UpdateFrequency.QUARTERLY,
            "is_active": True,
        },
    )
    return source


# ---------------------------------------------------------------------------
#  Чтение книг публикации
# ---------------------------------------------------------------------------


def read_series(spec: SeriesSpec, territories: dict[str, str]) -> dict[tuple[str, int], float]:
    """Прочитать один ряд публикации.

    Возвращает отображение ``(код территории, год) → значение в базовых
    единицах``. Разметка листа не задана жёстко: строка заголовка отыскивается
    по первой колонке с четырёхзначным годом, а строки территорий — по подписи
    без номера сноски. Привязка к номерам строк сломалась бы при первом же
    изменении состава публикации.
    """
    import openpyxl

    path = books_dir() / spec.file
    if not path.exists():
        raise RosstatError(
            f"Книга публикации {path} отсутствует. Книги Росстата входят "
            f"в поставку проекта: без них ряд загрузить неоткуда."
        )

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if spec.sheet not in workbook.sheetnames:
        raise RosstatError(
            f"В книге {spec.file} нет листа «{spec.sheet}»: "
            f"имеются {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[spec.sheet]
    rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()

    years = _header_years(rows)
    if not years:
        raise RosstatError(
            f"В листе «{spec.sheet}» книги {spec.file} не найдена строка "
            f"с годами: разметка публикации изменилась"
        )

    lookup = {label: code for code, label in territories.items()}
    series: dict[tuple[str, int], float] = {}
    for row in rows:
        label = strip_footnote(row[0] if row else "")
        code = lookup.get(label)
        if code is None:
            continue
        for column, year in years.items():
            if column >= len(row):
                continue
            value = row[column]
            if value is None or isinstance(value, str):
                # Прочерк и текстовое примечание означают отсутствие
                # наблюдения. Подставлять вместо них ноль нельзя: ноль —
                # это утверждение о том, что перевозок не было.
                continue
            series[(code, year)] = float(value) * spec.factor
    return series


def _header_years(rows: list[tuple]) -> dict[int, int]:
    """Отыскать строку заголовка и вернуть отображение «колонка → год»."""
    for row in rows[:20]:
        years = {}
        for index, cell in enumerate(row):
            if index == 0 or cell is None:
                continue
            match = _YEAR_RE.match(str(cell).strip())
            if match:
                years[index] = int(match.group(1))
        if len(years) >= 5:
            return years
    return {}


# ---------------------------------------------------------------------------
#  Конвейер
# ---------------------------------------------------------------------------


def _has_measure(candidate: Candidate) -> str | None:
    """Запись должна содержать хотя бы один показатель.

    Год, за который не опубликовано ни перевозок, ни грузооборота, записи
    не образует: пустая строка в ряду неотличима от нулевого значения
    и портит расчёт динамики.
    """
    if candidate.values.get("volume_tons") is None and (
        candidate.values.get("turnover_ton_km") is None
    ):
        return "за год не опубликовано ни одного показателя"
    return None


def _plausible_year(candidate: Candidate) -> str | None:
    """Год наблюдения должен быть правдоподобным."""
    year = candidate.extra.get("year")
    if year is None:
        return "год наблюдения не определён"
    if not FIRST_YEAR <= int(year) <= date.today().year + 1:
        return f"год наблюдения вне допустимых пределов: {year}"
    return None


class FreightStatisticsPipeline(ModelPipeline):
    """Годовые ряды перевозок и грузооборота автомобильного транспорта."""

    name = "rosstat.freight"
    title = "Статистика перевозок грузов"
    target_table = "freight_flow_stats"
    source_code = SOURCE_CODE
    description = (
        "Книги Росстата из data/reference/rosstat: перевозка грузов и "
        "грузооборот, по всем перевозчикам и на коммерческой основе, "
        "по Москве, области, округу и стране."
    )
    model = FreightFlowStat
    frequency = UpdateFrequency.QUARTERLY
    volatile_fields = ()
    checks: tuple[Check, ...] = (
        required("period_date", "Период наблюдения"),
        required("territory", "Территория"),
        Check("series.year", "Год наблюдения правдоподобен", _plausible_year),
        Check("series.measure", "Наблюдение содержит показатель", _has_measure),
        not_negative("volume_tons", "Объём перевозок"),
        not_negative("turnover_ton_km", "Грузооборот"),
        within("volume_tons", 0, 1e13, "Объём перевозок", "т"),
    )

    def ensure_source(self) -> DataSource:
        return ensure_source()

    def lookup(self, candidate: Candidate) -> dict:
        return {
            "source": candidate.extra["source"],
            "external_key": candidate.key,
        }

    def extract(self, context: Context) -> Extract:
        manifest = read(MANIFEST_FILE)
        territories = {
            item["code"]: item["label"] for item in manifest.payload.get("territories", [])
        }
        specs = [SeriesSpec(**item) for item in manifest.payload.get("series", [])]
        if not territories or not specs:
            raise RosstatError(
                f"Описание рядов {MANIFEST_FILE} не содержит ни территорий, "
                f"ни составов рядов"
            )

        observations: dict[tuple[str, int], Observation] = {}
        for spec in specs:
            for (code, year), value in read_series(spec, territories).items():
                key = (code, year, spec.scope)
                observation = observations.get(key)
                if observation is None:
                    observation = Observation(
                        territory_code=code,
                        territory=territories[code],
                        year=year,
                        values={},
                        sources={},
                    )
                    observations[key] = observation
                observation.values[spec.measure] = value
                observation.sources[spec.measure] = f"{spec.file}, лист {spec.sheet}"

        records = [
            (scope, observation)
            for (_code, _year, scope), observation in sorted(
                observations.items(), key=lambda item: (item[0][0], item[0][2], item[0][1])
            )
        ]
        return Extract(records=records, count=len(records), fetched_at=None)

    def prepare(self, extract: Extract, context: Context,
                report: RunReport) -> Iterator[Candidate]:
        source = self.ensure_source()

        for scope, observation in extract.records:
            volume = observation.values.get("volume_tons")
            turnover = observation.values.get("turnover_ton_km")
            yield Candidate(
                key=f"rosstat:{scope}:{observation.key}",
                position=f"{observation.territory}, {observation.year}",
                values={
                    "period_date": date(observation.year, 1, 1),
                    "period_type": PeriodType.YEAR,
                    "territory": observation.territory,
                    "direction": FlowDirection.TOTAL,
                    "scope": scope,
                    "volume_tons": round(volume, 2) if volume is not None else None,
                    "turnover_ton_km": round(turnover, 2) if turnover is not None else None,
                    "origin": DataOrigin.MEASURED,
                    "source": source,
                },
                extra={"source": source, "year": observation.year, "scope": scope},
                payload={
                    "территория": observation.territory,
                    "год": observation.year,
                    "круг перевозчиков": scope,
                    "показатели": observation.values,
                    "листы": observation.sources,
                },
            )

    def verify(self, report: RunReport, context: Context) -> None:
        """Сообщить глубину рядов и полноту наблюдений.

        Глубина ряда определяет, какие модели прогноза вообще применимы:
        ряд короче десятка наблюдений не позволяет ни оценить сезонность,
        ни выделить отложенную выборку. Величина эта важнее числа записей
        и потому выводится отдельно.
        """
        source = ensure_source()
        rows = FreightFlowStat.objects.filter(source=source)
        for scope, label in FlowScope.choices:
            # Порядок по умолчанию попадает в SELECT и делает выборку
            # неразличимой: перечень территорий приходится запрашивать
            # без сортировки.
            territories = (
                rows.filter(scope=scope)
                .order_by()
                .values_list("territory", flat=True)
                .distinct()
            )
            for territory in sorted(territories):
                series = rows.filter(scope=scope, territory=territory)
                years = sorted(series.values_list("period_date__year", flat=True))
                if not years:
                    continue
                gaps = [
                    year for year in range(years[0], years[-1] + 1) if year not in years
                ]
                report.detail(
                    f"{territory}, {label.lower()}: {len(years)} наблюдений "
                    f"за {years[0]}–{years[-1]}"
                )
                if gaps:
                    report.note(
                        f"{territory}, {label.lower()}: пропуски в ряду — "
                        f"{', '.join(str(year) for year in gaps)}"
                    )


__all__ = [
    "FreightStatisticsPipeline",
    "MANIFEST_FILE",
    "Observation",
    "RosstatError",
    "SeriesSpec",
    "read_series",
    "strip_footnote",
]
