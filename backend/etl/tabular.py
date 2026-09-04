"""Чтение табличных выгрузок: CSV и книги Excel.

Пользователь присылает ряд в том виде, в каком он у него есть: выгрузкой
из учётной системы, книгой Excel, файлом с разделителями. Приводить такие
файлы к единому виду вручную бессмысленно — это работа, которую и должна
выполнять система.

Модуль отвечает только за чтение: превращает файл в строки со словарями
значений и с номерами строк. Смысл колонок, проверки и запись — дело
конвейера, который эту таблицу получает. Номер строки сохраняется для
каждой записи: без него отклонённую строку невозможно отыскать в присланном
файле и исправить.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger("freightflow.etl")

#: Кодировки, в которых приходят файлы. Порядок значим: разбор в utf-8
#: на файле в кодировке Windows отказывает, обратное же молча даёт
#: нечитаемые подписи колонок.
ENCODINGS = ("utf-8-sig", "utf-8", "cp1251")

#: Разделители, встречающиеся в выгрузках. Точка с запятой — обычный
#: разделитель для Excel в русской локали, где запятая занята дробной частью.
DELIMITERS = ";,\t"

#: Наибольшее число строк в присланном файле. Ограничение защищает от
#: выгрузки, присланной по ошибке: ряд статистики — это сотни строк,
#: а не миллионы.
MAX_ROWS = 50_000

#: Расширения, распознаваемые как книги Excel.
WORKBOOK_SUFFIXES = (".xlsx", ".xlsm")


class TableError(ValueError):
    """Файл не читается как таблица."""


@dataclass(frozen=True)
class TableRow:
    """Строка таблицы вместе с её номером в файле."""

    number: int
    values: dict[str, str]

    def get(self, column: str) -> str:
        return (self.values.get(column) or "").strip()

    @property
    def is_empty(self) -> bool:
        return not any(value and str(value).strip() for value in self.values.values())


@dataclass
class Table:
    """Прочитанная таблица."""

    name: str
    columns: list[str] = field(default_factory=list)
    rows: list[TableRow] = field(default_factory=list)

    def __len__(self) -> int:
        return len(self.rows)


def normalize_header(value: object) -> str:
    """Привести подпись колонки к виду, пригодному для сопоставления.

    Регистр, лишние пробелы и знаки препинания в подписях колонок различаются
    от выгрузки к выгрузке, а обозначают одно и то же. Единицы измерения
    в скобках сохраняются: «объём, т» и «объём, тыс. т» — разные колонки.
    """
    text = str(value or "").strip().lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .:")


def read_table(source: Path | str | bytes, name: str = "") -> Table:
    """Прочитать таблицу из файла либо из его содержимого.

    Формат определяется по расширению имени: книга Excel разбирается
    средствами openpyxl, всё остальное читается как текст с разделителями.
    """
    title = name or (Path(source).name if isinstance(source, str | Path) else "выгрузка")
    if title.lower().endswith(WORKBOOK_SUFFIXES):
        return _read_workbook(source, title)
    return _read_delimited(source, title)


# ---------------------------------------------------------------------------
#  Текст с разделителями
# ---------------------------------------------------------------------------


def _read_delimited(source: Path | str | bytes, name: str) -> Table:
    raw = source if isinstance(source, bytes) else Path(source).read_bytes()
    text = _decode(raw, name)
    delimiter = _sniff(text)

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        header = next(reader)
    except StopIteration:
        raise TableError(f"Файл {name} пуст") from None

    columns = [normalize_header(cell) for cell in header]
    _require_columns(columns, name)

    table = Table(name=name, columns=columns)
    # Нумерация строк ведётся от единицы и включает заголовок: именно такой
    # номер пользователь видит в своём редакторе.
    for number, row in enumerate(reader, start=2):
        if number - 1 > MAX_ROWS:
            raise TableError(
                f"В файле {name} больше {MAX_ROWS} строк: такая выгрузка "
                f"загружается через командную строку, а не через форму"
            )
        values = {
            column: (row[index] if index < len(row) else "")
            for index, column in enumerate(columns)
        }
        entry = TableRow(number=number, values=values)
        if not entry.is_empty:
            table.rows.append(entry)
    return table


def _decode(raw: bytes, name: str) -> str:
    for encoding in ENCODINGS:
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise TableError(
        f"Кодировку файла {name} определить не удалось: "
        f"проверены {', '.join(ENCODINGS)}"
    )


def _sniff(text: str) -> str:
    """Определить разделитель по первой строке файла.

    Встроенный определитель csv на строке из одной колонки ошибается
    и выбирает случайный знак, поэтому подсчёт ведётся напрямую: побеждает
    тот разделитель, который встречается в заголовке чаще.
    """
    head = text.splitlines()[0] if text else ""
    counts = {delimiter: head.count(delimiter) for delimiter in DELIMITERS}
    best = max(counts, key=lambda key: counts[key])
    return best if counts[best] else ";"


# ---------------------------------------------------------------------------
#  Книги Excel
# ---------------------------------------------------------------------------


def _read_workbook(source: Path | str | bytes, name: str) -> Table:
    import openpyxl

    handle = io.BytesIO(source) if isinstance(source, bytes) else source
    try:
        workbook = openpyxl.load_workbook(handle, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — сообщение отдаётся пользователю
        raise TableError(f"Книга {name} не читается: {exc}") from exc

    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)

    columns: list[str] = []
    table = Table(name=name)
    for number, row in enumerate(rows, start=1):
        if not columns:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            columns = [normalize_header(cell) for cell in row]
            _require_columns(columns, name)
            table.columns = columns
            continue

        if number - 1 > MAX_ROWS:
            workbook.close()
            raise TableError(
                f"В книге {name} больше {MAX_ROWS} строк: такая выгрузка "
                f"загружается через командную строку, а не через форму"
            )
        values = {
            column: ("" if cell is None else str(cell))
            for column, cell in zip(columns, row, strict=False)
        }
        entry = TableRow(number=number, values=values)
        if not entry.is_empty:
            table.rows.append(entry)

    workbook.close()
    if not columns:
        raise TableError(f"В книге {name} не найдена строка заголовка")
    return table


def _require_columns(columns: list[str], name: str) -> None:
    if not any(columns):
        raise TableError(f"В файле {name} не распознана строка заголовка")


__all__ = ["MAX_ROWS", "Table", "TableError", "TableRow", "normalize_header", "read_table"]
